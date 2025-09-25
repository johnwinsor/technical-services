#!/usr/bin/env python3
"""
Workday Excel to EDIFACT EDI Converter
Converts Workday Supplier Invoice Excel exports to EDIFACT INVOIC format for Alma
"""

import datetime
from typing import Dict, List, Any, Optional
from dataclasses import dataclass
import re
import sys
import os


@dataclass
class EDIFACTConfig:
    """Configuration for EDIFACT generation"""
    sender_id: str = "1694510101"  # Amazon vendor ID in your system
    sender_qualifier: str = "31B"
    receiver_id: str = "3333159"  # From your sample  
    receiver_qualifier: str = "31B"
    interchange_ref: str = None  # Will be generated
    currency: str = "USD"
    vendor_account: str = "amazon"  # Always "amazon" for Amazon Business orders


class EDIFACTInvoiceGenerator:
    """Generates EDIFACT INVOIC messages from Workday data"""
    
    def __init__(self, config: EDIFACTConfig):
        self.config = config
        if not self.config.interchange_ref:
            self.config.interchange_ref = datetime.datetime.now().strftime("%m%d%H%M%S%f")[:-3]
    
    def generate_una_unb_headers(self) -> str:
        """Generate UNA and UNB headers"""
        now = datetime.datetime.now()
        date_str = now.strftime("%y%m%d")
        time_str = now.strftime("%H%M")
        
        una = "UNA:+.? '"
        unb = (f"UNB+UNOC:2+{self.config.sender_id}:{self.config.sender_qualifier}+"
               f"{self.config.receiver_id}:{self.config.receiver_qualifier}+"
               f"{date_str}:{time_str}+{self.config.interchange_ref}'")
        
        return una + unb
    
    def generate_invoice_segments(self, invoice_data: Dict[str, Any]) -> List[str]:
        """Generate all segments for a single invoice"""
        segments = []
        
        # UNH - Message Header
        msg_ref = invoice_data.get('invoice_number', '1')
        segments.append(f"UNH+{msg_ref}+INVOIC:D:96A:UN:EAN008'")
        
        # BGM - Beginning of Message
        invoice_num = invoice_data.get('invoice_number', '491150')
        segments.append(f"BGM+380+{invoice_num}'")
        
        # DTM - Date/Time (Invoice Date)
        invoice_date = invoice_data.get('invoice_date', datetime.datetime.now().strftime("%Y%m%d"))
        segments.append(f"DTM+137:{invoice_date}:102'")
        
        # DTM - Date/Time (Due Date) - if available
        due_date = invoice_data.get('due_date', '')
        if due_date and due_date != invoice_date:
            segments.append(f"DTM+13:{due_date}:102'")
        
        # RFF - Reference (API reference - using Family or default)
        api_ref = invoice_data.get('api_ref', '7015-10')
        segments.append(f"RFF+API:{api_ref}'")
        
        # RFF - Vendor Account Reference (from config)
        segments.append(f"RFF+VA:{self.config.vendor_account}'")
        
        # CUX - Currencies
        segments.append(f"CUX+2:{self.config.currency}:4'")
        
        # Calculate charges
        line_items = invoice_data.get('line_items', [])
        shipping_total = sum(item.get('shipping', 0) for item in line_items)
        total_tax = invoice_data.get('total_tax', 0)
        
        # ALC - Allowance or Charge (Freight) - only if there's shipping cost
        if shipping_total > 0:
            segments.append("ALC+C++++DL::28:Freight Charges'")
            segments.append(f"MOA+8:{shipping_total:.2f}'")
        
        # ALC - Allowance or Charge (Tax) - if there's tax
        if total_tax > 0:
            segments.append("ALC+C++++TX::28:Sales Tax'")
            segments.append(f"MOA+8:{total_tax:.2f}'")
        
        # Line items
        for i, item in enumerate(line_items, 1):
            segments.extend(self.generate_line_item_segments(item, i))
        
        # UNS - Section Control
        segments.append("UNS+S'")
        
        # CNT - Control Total
        total_lines = len(line_items)
        total_qty = sum(int(item.get('quantity', 1) or 1) for item in line_items)
        segments.append(f"CNT+1:{total_qty}'")
        segments.append(f"CNT+2:{total_lines}'")
        
        # MOA - Monetary Amount
        gross_total = sum(item.get('unit_price', 0) * int(item.get('quantity', 1) or 1) 
                         for item in line_items) + shipping_total
        
        # Get total tax from invoice info (Workday calculates this)
        total_tax = invoice_data.get('total_tax', 0)
        net_total = invoice_data.get('total_amount', gross_total + total_tax)
        
        segments.append(f"MOA+9:{gross_total:.2f}'")
        segments.append(f"MOA+79:{net_total:.2f}'")
        
        # Note: Tax is now handled as ALC charge above, not as separate MOA+176
        
        # UNT - Message Trailer
        segment_count = len(segments) + 1  # +1 for the UNT segment itself
        segments.append(f"UNT+{segment_count}+{msg_ref}'")
        
        return segments
    
    def generate_line_item_segments(self, item: Dict[str, Any], line_num: int) -> List[str]:
        """Generate segments for a single line item"""
        segments = []
        
        # LIN - Line Item
        item_id = item.get('supplier_item_id', '')
        segments.append(f"LIN+{line_num}++{item_id}:EN'")
        
        # IMD - Item Description - Author (extracted from title if available)
        author = item.get('author', '').upper()
        if author:
            segments.append(f"IMD+L+010+:::{author}'")
        
        # IMD - Item Description - Title (may need multiple segments if long)
        title = item.get('title', '').upper()
        title_segments = self.split_title_for_imd(title)
        for title_seg in title_segments:
            segments.append(f"IMD+L+050+:::{title_seg}'")
        
        # QTY - Quantity
        quantity = item.get('quantity', '1') or '1'
        segments.append(f"QTY+47:{quantity}'")
        
        # MOA - Monetary Amount (line total)
        unit_price = item.get('unit_price', 0)
        line_total = unit_price * int(quantity)
        segments.append(f"MOA+203:{line_total:.2f}'")
        
        # PRI - Price Details
        list_price = item.get('list_price', unit_price)
        segments.append(f"PRI+AAB:{list_price:.2f}'")
        segments.append(f"PRI+AAA:{unit_price:.2f}'")
        
        # TAX - Duty/tax/fee details (if tax information available)
        tax_rate = item.get('tax_rate', '')
        tax_amount = item.get('tax_amount', 0)
        if tax_rate and tax_rate != '0':
            segments.append(f"TAX+7+VAT+++:::{tax_rate}'")
            if tax_amount > 0:
                segments.append(f"MOA+124:{tax_amount:.2f}'")
        
        # RFF - References
        pol = item.get('pol', '')
        if pol:
            segments.append(f"RFF+LI:{pol}'")
        
        po_line_ref = item.get('po_line_ref', '0')
        segments.append(f"RFF+SLI:{po_line_ref}'")
        
        return segments
    
    def split_title_for_imd(self, title: str, max_length: int = 35) -> List[str]:
        """Split long titles into multiple IMD segments"""
        if len(title) <= max_length:
            return [title]
        
        # Try to split on word boundaries
        words = title.split()
        segments = []
        current_segment = ""
        
        for word in words:
            if len(current_segment + " " + word) <= max_length:
                if current_segment:
                    current_segment += " " + word
                else:
                    current_segment = word
            else:
                if current_segment:
                    segments.append(current_segment)
                    current_segment = word
                else:
                    # Word is longer than max_length, force split
                    segments.append(word[:max_length])
                    current_segment = word[max_length:]
        
        if current_segment:
            segments.append(current_segment)
        
        return segments
    
    def generate_unz_trailer(self, message_count: int) -> str:
        """Generate UNZ trailer"""
        return f"UNZ+{message_count}+{self.config.interchange_ref}'"


def parse_workday_excel(excel_file_path: str) -> List[Dict[str, Any]]:
    """Parse Workday Excel file and extract invoice data"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required to read Excel files. Install it with: pip install openpyxl")
    
    # Load the workbook
    workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
    
    # Use the "Invoice Lines" sheet which has the main data
    if "Invoice Lines" not in workbook.sheetnames:
        raise ValueError("Expected 'Invoice Lines' sheet not found in Excel file")
    
    sheet = workbook["Invoice Lines"]
    
    # Extract header information (key-value pairs in columns A and B)
    invoice_info = {}
    
    # Read through the sheet to find key information
    for row in range(1, 50):  # Check first 50 rows for header info
        key_cell = sheet.cell(row=row, column=1).value
        value_cell = sheet.cell(row=row, column=2).value
        
        if key_cell and isinstance(key_cell, str):
            key_lower = key_cell.lower()
            if "invoice number" in key_lower and value_cell:
                invoice_info['invoice_number'] = str(value_cell)
            elif "supplier's invoice number" in key_lower and value_cell:
                invoice_info['suppliers_invoice_number'] = str(value_cell)
            elif "invoice date" in key_lower and value_cell:
                if isinstance(value_cell, datetime.datetime):
                    invoice_info['invoice_date'] = value_cell.strftime("%Y%m%d")
                else:
                    # Try parsing the string value
                    parsed_date = parse_date(str(value_cell))
                    if parsed_date != datetime.datetime.now().strftime("%Y%m%d"):  # Only use if parsing succeeded
                        invoice_info['invoice_date'] = parsed_date
            elif "due date" in key_lower and value_cell:
                if isinstance(value_cell, datetime.datetime):
                    invoice_info['due_date'] = value_cell.strftime("%Y%m%d")
                else:
                    # Try parsing the string value
                    parsed_date = parse_date(str(value_cell))
                    if parsed_date != datetime.datetime.now().strftime("%Y%m%d"):  # Only use if parsing succeeded
                        invoice_info['due_date'] = parsed_date
            elif "total invoice amount" in key_lower and value_cell:
                invoice_info['total_amount'] = float(value_cell)
            elif "tax amount" in key_lower and value_cell:
                invoice_info['total_tax'] = float(value_cell)
            elif "currency" in key_lower and value_cell:
                invoice_info['currency'] = str(value_cell)
    
    # Find the line items table (starts around row 41-42 based on the sample)
    line_items = []
    header_row = None
    spend_category = None  # Will be used for RFF+API
    
    # Look for the line items header row
    for row in range(35, 50):
        cell_value = sheet.cell(row=row, column=1).value
        if cell_value and "Invoice Line" in str(cell_value) and "Company" in str(sheet.cell(row=row, column=2).value or ""):
            header_row = row
            break
    
    if header_row:
        # Get column headers
        headers = {}
        for col in range(1, 50):  # Check first 50 columns
            header_value = sheet.cell(row=header_row, column=col).value
            if header_value:
                headers[col] = str(header_value)
        
        # Read line item data
        for row in range(header_row + 1, sheet.max_row + 1):
            # Check if this row has line item data
            first_cell = sheet.cell(row=row, column=1).value
            if not first_cell or not str(first_cell).startswith("Supplier Invoice:"):
                continue
            
            line_item = {}
            
            # Extract data from each column
            for col, header in headers.items():
                cell_value = sheet.cell(row=row, column=col).value
                
                if "Line Item Description" in header and cell_value:
                    line_item['title'] = clean_title(str(cell_value))
                elif "Supplier Item Identifier" in header and cell_value:
                    line_item['supplier_item_id'] = str(cell_value)
                elif "Business Document" in header and cell_value:
                    line_item['po_line_ref'] = extract_po_line_ref(str(cell_value))
                elif "Spend Category" in header and cell_value:
                    # Use spend category for RFF+API - should be same for all lines
                    if spend_category is None:
                        spend_category = str(cell_value)
                elif "Quantity" in header and cell_value:
                    line_item['quantity'] = int(cell_value) if cell_value else 1
                elif "Unit Cost" in header and cell_value:
                    line_item['unit_price'] = float(cell_value)
                    line_item['list_price'] = float(cell_value)  # Same as unit cost for Workday
                elif "Extended Amount" in header and cell_value:
                    line_item['extended_amount'] = float(cell_value)
                elif "POL" in header and cell_value:  # Manual POL column
                    line_item['pol'] = str(cell_value).strip()
                elif header == "POL" and cell_value:  # Exact match for POL column
                    line_item['pol'] = str(cell_value).strip()
            
            # Extract author from title if possible
            if 'title' in line_item:
                line_item['author'] = extract_author_from_title(line_item['title'])
                line_item['title'] = clean_title_remove_author(line_item['title'], line_item['author'])
            
            line_items.append(line_item)
    
    # Create the invoice record
    invoice = {
        'invoice_number': invoice_info.get('invoice_number', 'UNKNOWN'),
        'invoice_date': invoice_info.get('invoice_date', datetime.datetime.now().strftime("%Y%m%d")),
        'due_date': invoice_info.get('due_date', ''),
        'api_ref': spend_category if spend_category else '7015-10',  # Use spend category or fallback
        'total_amount': invoice_info.get('total_amount', 0),
        'total_tax': invoice_info.get('total_tax', 0),
        'line_items': line_items
    }
    
    # Warn if invoice date looks suspicious
    if invoice['invoice_date'] == '19691231':
        print("⚠️  Warning: Invoice date shows as 1969 - check Excel date formatting")
    
    return [invoice]  # Return as list for consistency with Amazon script


def extract_po_line_ref(business_doc: str) -> str:
    """Extract PO line reference from business document field"""
    # Pattern like "PO00008234 - Line 1"
    match = re.search(r'(PO\d+\s*-\s*Line\s*\d+)', business_doc)
    if match:
        return match.group(1).replace(" ", "")
    return business_doc[:20]  # Fallback to first 20 chars


def extract_pol_from_po_ref(business_doc: str) -> str:
    """Extract POL reference from business document field"""
    # POL references need to be manually added to the Excel file
    # This function is kept for potential future use but returns empty
    return ""  # No POL found in Workday data - must be added manually


def clean_title(title: str) -> str:
    """Clean up title text"""
    # Remove extra whitespace and normalize
    title = re.sub(r'\s+', ' ', title.strip())
    return title


def extract_author_from_title(title: str) -> str:
    """Extract author name from title using common patterns"""
    # This is a simple heuristic - adjust based on your data
    if ' by ' in title.lower():
        parts = title.split(' by ')
        if len(parts) > 1:
            return parts[-1].strip()
    
    # Look for patterns like "Author Name - Title"
    if ' - ' in title:
        parts = title.split(' - ')
        if len(parts) > 1 and parts[0].replace(' ', '').replace('.', '').isalpha():
            return parts[0].strip()
    
    return ""


def clean_title_remove_author(title: str, author: str) -> str:
    """Remove author name from title if it was extracted"""
    if author and author in title:
        title = title.replace(f" by {author}", "").replace(f"{author} - ", "").strip()
    return title


def parse_date(date_str: str) -> str:
    """Parse various date formats to YYYYMMDD"""
    if not date_str or date_str.strip() == '':
        return datetime.datetime.now().strftime("%Y%m%d")
    
    # Clean up the date string
    date_str = str(date_str).strip()
    
    # Skip obviously bad dates (Unix epoch, etc.)
    if '1969' in date_str or '1970' in date_str:
        print(f"⚠️  Warning: Skipping suspicious date: {date_str}")
        return datetime.datetime.now().strftime("%Y%m%d")
    
    # Try common formats
    formats = [
        '%m/%d/%Y', '%Y-%m-%d', '%m-%d-%Y', '%Y/%m/%d', 
        '%B %d, %Y', '%b %d, %Y', '%Y%m%d',
        '%m/%d/%y', '%Y-%m-%d %H:%M:%S'
    ]
    
    for fmt in formats:
        try:
            date_obj = datetime.datetime.strptime(date_str, fmt)
            # Additional validation - reject dates before 2000 or far in future
            if date_obj.year < 2000 or date_obj.year > 2030:
                continue
            return date_obj.strftime("%Y%m%d")
        except ValueError:
            continue
    
    print(f"⚠️  Warning: Could not parse date '{date_str}', using current date")
    # If all else fails, return current date
    return datetime.datetime.now().strftime("%Y%m%d")


def convert_workday_excel_to_edifact(excel_file_path: str, edi_file_path: str, config: EDIFACTConfig = None):
    """Convert Workday Excel to EDIFACT EDI format"""
    if config is None:
        config = EDIFACTConfig()
    
    # Parse Excel
    invoices = parse_workday_excel(excel_file_path)
    
    if not invoices:
        raise ValueError("No invoice data found in Excel file")
    
    # Print summary of what will be created
    print(f"\n📋 Processing {len(invoices)} invoice(s) from Workday:")
    print("=" * 60)
    
    total_line_items = 0
    for invoice in invoices:
        line_count = len(invoice['line_items'])
        total_line_items += line_count
        total_amount = invoice.get('total_amount', 0)
        print(f"Invoice: {invoice['invoice_number']} → {line_count} line item(s) → ${total_amount:.2f}")
        
        # Show line items for verification
        for i, item in enumerate(invoice['line_items'], 1):
            title = item.get('title', 'Unknown Item')[:50] + ('...' if len(item.get('title', '')) > 50 else '')
            price = item.get('unit_price', 0)
            print(f"  {i}. {title} → ${price:.2f}")
    
    print("=" * 60)
    print(f"📊 Total: {len(invoices)} invoices, {total_line_items} line items")
    print()
    
    # Generate EDI
    generator = EDIFACTInvoiceGenerator(config)
    
    with open(edi_file_path, 'w', newline='', encoding='utf-8') as edi_file:
        # Write headers
        edi_file.write(generator.generate_una_unb_headers())
        
        # Write each invoice
        for invoice in invoices:
            segments = generator.generate_invoice_segments(invoice)
            for segment in segments:
                edi_file.write(segment)
        
        # Write trailer
        edi_file.write(generator.generate_unz_trailer(len(invoices)))


def main():
    """Main function with command line argument support"""
    # Check for command line argument
    if len(sys.argv) != 2:
        print("Usage: python workday_to_edi.py <excel_file_path>")
        print("Example: python workday_to_edi.py /path/to/View_Supplier_Invoice.xlsx")
        sys.exit(1)
    
    excel_file_path = sys.argv[1]
    
    # Check if Excel file exists
    if not os.path.exists(excel_file_path):
        print(f"❌ Error: Excel file '{excel_file_path}' not found.")
        sys.exit(1)
    
    # Generate EDI file path in same directory as Excel file
    excel_dir = os.path.dirname(excel_file_path)
    excel_filename = os.path.basename(excel_file_path)
    excel_name, _ = os.path.splitext(excel_filename)
    edi_file_path = os.path.join(excel_dir, f"{excel_name}.edi")
    
    # Configuration - uses defaults from EDIFACTConfig class
    config = EDIFACTConfig()
    
    try:
        convert_workday_excel_to_edifact(excel_file_path, edi_file_path, config)
        print(f"✅ Successfully converted '{excel_file_path}' to EDIFACT EDI format!")
        print(f"📁 Output file: '{edi_file_path}'")
        print("💡 Note: Make sure you've added a 'POL' column to your Excel file with Alma POL numbers")
    except ImportError as e:
        print(f"❌ Error: {e}")
        print("💡 Install required dependencies with: pip install openpyxl")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Error during conversion: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()